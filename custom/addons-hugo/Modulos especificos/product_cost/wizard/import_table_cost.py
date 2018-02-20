# -*- coding: utf-8 -*-
from openerp import models, fields, exceptions, api, _

import base64
import cStringIO
import csv
import numbers


class ImportTableCost(models.TransientModel):
    _name = 'import.table.cost'
    _description = 'Import Table Cost'
    
    data = fields.Binary(
        string='File', required=True)
    name = fields.Char(
        string='Filename', required=False)
    delimeter = fields.Char(
        string='Delimeter', default=',',
        help='Default delimeter is ","')
    file_type = fields.Selection(
        selection=[#('csv', 'CSV'), 
                   ('xls', 'XLS'),
                   ('xlsx', 'XLSX')],
        string='File Type', required=True,
        default='xls')
    
    def _import_csv(self, load_id, file_data, mode2d, delimeter=';'):
        """ Imports data from a CSV file in defined object.
        @param load_id: Loading id
        @param file_data: Input data to load
        @param delimeter: CSV file data delimeter
        @return: Imported file number
        """
        return 0 #TODO
        file_line_obj = self.env['product.pricelist.load.line']
        data = base64.b64decode(file_data)
        file_input = cStringIO.StringIO(data)
        file_input.seek(0)
        reader_info = []
        reader = csv.reader(file_input, delimiter=str(delimeter),
                            lineterminator='\r\n')
        try:
            reader_info.extend(reader)
        except Exception:
            raise exceptions.Warning(_("Not a valid file!"))
        # keys2 = reader_info[0]
        counter = 0
        keys = ['code', 'info', 'price', 'discount_1', 'discount_2',
                'retail', 'pdv1', 'pdv2']
        if not isinstance(keys, list):
            raise exceptions.Warning(_("Not a valid file!"))
        del reader_info[0]
        for i in range(len(reader_info)):
            field = reader_info[i]
            values = dict(zip(keys, field))
            file_line_obj.create(
                {'code': values['code'], 'info': values['info'],
                 'price': values['price'].replace(',', '.'),
                 'discount_1': float(values['discount_1'].replace(',', '.')),
                 'discount_2': float(values['discount_2'].replace(',', '.')),
                 'retail': float(values['retail'].replace(',', '.')),
                 'pdv1': float(values['pdv1'].replace(',', '.')),
                 'pdv2': float(values['pdv2'].replace(',', '.')),
                 'fail': True, 'fail_reason': _('No Processed'),
                 'file_load': load_id
                 })
            counter += 1
        return counter

    def _import_xls2d(self, template_id, file_data):
        """ Imports data from a XLS file in defined object.
        @param load_id: Loading id
        @param file_data: Input data to load
        @return: Imported file number
        """
        try:
            import xlrd
        except ImportError:
            exceptions.Warning(_("xlrd python lib  not installed"))
        table_cost_item_obj = self.env['template.table.cost.item']
        file_1 = base64.decodestring(file_data)
        book = xlrd.open_workbook('dummy_name.xls', file_contents=file_1)
        sheet = book.sheet_by_index(0)
        
        counter = 0
        for row_index in range(1, sheet.nrows):
            for col_index in range(1, sheet.ncols):
                cell = sheet.cell(row_index, col_index)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                
                if not isinstance(cell.value, numbers.Number):
                    raise exceptions.Warning(_('Found non numeric value in a non-empty cell. Only numbers allowed.'))
                
                x_lower = sheet.cell_value(0, col_index-1) if col_index > 1 else 0.0
                x_upper = sheet.cell_value(0, col_index)
                if (not isinstance(x_lower, numbers.Number)) or (not isinstance(x_upper, numbers.Number)):
                    raise exceptions.Warning(_('Non numeric value found as X limit'))
                
                y_lower = sheet.cell_value(row_index-1, 0) if row_index > 1 else 0.0
                y_upper = sheet.cell_value(row_index, 0)
                if (not isinstance(y_lower, numbers.Number)) or (not isinstance(y_upper, numbers.Number)):
                    raise exceptions.Warning(_('Non numeric value found as Y limit'))
                
                table_cost_item_obj.create({
                    'template_id': template_id,
                    'x_upper': x_upper,
                    'x_lower': x_lower,
                    'y_upper': y_upper,
                    'y_lower': y_lower,
                    'cost': cell.value,})
                counter += 1
        
        return counter

    def _import_xlsx2d(self, template_id, file_data):
        """ Imports data from a XLSX file in defined object.
        @param load_id: Loading id
        @param file_data: Input data to load
        @return: Imported file number
        """
        try:
            import openpyxl
        except ImportError:
            exceptions.Warning(_("openpyxl python lib  not installed"))
        table_cost_item_obj = self.env['template.table.cost.item']
        data = base64.decodestring(file_data)
        file_input = cStringIO.StringIO(data)
        file_input.seek(0)
        book = openpyxl.load_workbook(filename=file_input)
        sheet = book.worksheets[0]
        
        counter = 0
        for row_index in range(1, sheet.get_highest_row()):
            for col_index in range(1, sheet.get_highest_column()):
                cell = sheet.cell(row=row_index, column=col_index)
                if cell.data_type == openpyxl.cell.Cell.TYPE_NULL:
                    continue
                
                if not isinstance(cell.value, numbers.Number):
                    raise exceptions.Warning(_('Found non numeric value in a non-empty cell. Only numbers allowed.'))
                
                x_lower = sheet.cell(row=0, column=col_index-1).value if col_index > 1 else 0.0
                x_upper = sheet.cell(row=0, column=col_index).value
                if (not isinstance(x_lower, numbers.Number)) or (not isinstance(x_upper, numbers.Number)):
                    raise exceptions.Warning(_('Non numeric value found as X limit'))
                
                y_lower = sheet.cell(row=row_index-1, column=0).value if row_index > 1 else 0.0
                y_upper = sheet.cell(row=row_index, column=0).value
                if (not isinstance(y_lower, numbers.Number)) or (not isinstance(y_upper, numbers.Number)):
                    raise exceptions.Warning(_('Non numeric value found as Y limit'))
                
                table_cost_item_obj.create({
                    'template_id': template_id,
                    'x_upper': x_upper,
                    'x_lower': x_lower,
                    'y_upper': y_upper,
                    'y_lower': y_lower,
                    'cost': cell.value,})
                counter += 1
        
        return counter

    def _import_xls1d(self, template_id, file_data):
        """ Imports data from a XLS file in defined object.
        @param load_id: Loading id
        @param file_data: Input data to load
        @return: Imported file number
        """
        try:
            import xlrd
        except ImportError:
            exceptions.Warning(_("xlrd python lib  not installed"))
        table_cost_item_obj = self.env['template.table.cost.item.one']
        file_1 = base64.decodestring(file_data)
        book = xlrd.open_workbook('dummy_name.xls', file_contents=file_1)
        sheet = book.sheet_by_index(0)
        
        counter = 0
        for col_index in range(0, sheet.ncols):
            cell = sheet.cell(1, col_index)
            if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                continue
            
            if not isinstance(cell.value, numbers.Number):
                raise exceptions.Warning(_('Found non numeric value in a non-empty cell. Only numbers allowed.'))
            
            x_lower = sheet.cell_value(0, col_index-1) if col_index >= 1 else 0.0
            x_upper = sheet.cell_value(0, col_index)
            if (not isinstance(x_lower, numbers.Number)) or (not isinstance(x_upper, numbers.Number)):
                raise exceptions.Warning(_('Non numeric value found as X limit'))
            
            table_cost_item_obj.create({
                'template_id': template_id,
                'x_upper': x_upper,
                'x_lower': x_lower,
                'cost': cell.value,})
            counter += 1
        
        return counter

    def _import_xlsx1d(self, template_id, file_data):
        """ Imports data from a XLSX file in defined object.
        @param load_id: Loading id
        @param file_data: Input data to load
        @return: Imported file number
        """
        try:
            import openpyxl
        except ImportError:
            exceptions.Warning(_("openpyxl python lib  not installed"))
        table_cost_item_obj = self.env['template.table.cost.item.one']
        data = base64.decodestring(file_data)
        file_input = cStringIO.StringIO(data)
        file_input.seek(0)
        book = openpyxl.load_workbook(filename=file_input)
        sheet = book.worksheets[0]
        
        counter = 0
        for col_index in range(0, sheet.get_highest_col()):
            cell = sheet.cell(row=1, column=col_index)
            if cell.data_type == openpyxl.cell.Cell.TYPE_NULL:
                continue
            
            if not isinstance(cell.value, numbers.Number):
                raise exceptions.Warning(_('Found non numeric value in a non-empty cell. Only numbers allowed.'))
            
            x_lower = sheet.cell(row=0, column=col_index-1).value if col_index > 1 else 0.0
            x_upper = sheet.cell(row=0, column=col_index).value
            if (not isinstance(x_lower, numbers.Number)) or (not isinstance(x_upper, numbers.Number)):
                raise exceptions.Warning(_('Non numeric value found as X limit'))
            
            table_cost_item_obj.create({
                'template_id': template_id,
                'x_upper': x_upper,
                'x_lower': x_lower,
                'cost': cell.value,})
            counter += 1
        
        return counter
    
    @api.multi
    def action_import(self):
        mode = self._context.get('table_dimension_mode')
        if not mode:
            raise exceptions.Warning(_("Could not find the table dimension mode.")) #should never happen
        
        if mode == 'table1d':
            self.action_import1d()
        elif mode == 'table2d':
            self.action_import2d()
        else:
            raise exceptions.Warning(_("Invalid dimension mode.")) #should never happen
    
    @api.multi
    def action_import2d(self):
        template = None
        if 'active_id' in self._context:
            template_id = self._context['active_id']
            template = self.env['product.template'].browse(template_id)
        if not template:
            raise exceptions.Warning(_("Could not find the product.")) #should never happen
           
        template.table_cost_items.unlink()
        
        for wiz in self:
            if not wiz.data:
                raise exceptions.Warning(_("You need to select a file!"))
            date_hour = fields.datetime.now()
            actual_date = fields.date.today()
            filename = wiz.name
            if wiz.file_type == 'csv':
                counter = self._import_csv2d(template.id, wiz.data, wiz.delimeter)
            elif wiz.file_type == 'xls':
                counter = self._import_xls2d(template.id, wiz.data)
            elif wiz.file_type == 'xlsx':
                counter = self._import_xlsx2d(template.id, wiz.data)
            else:
                raise exceptions.Warning(_("Not a .csv/.xls/.xlsx file found"))
        return counter

    @api.multi
    def action_import1d(self):
        template = None
        if 'active_id' in self._context:
            template_id = self._context['active_id']
            template = self.env['product.template'].browse(template_id)
        if not template:
            raise exceptions.Warning(_("Could not find the product.")) #should never happen
           
        template.table_cost_items1d.unlink()
        
        for wiz in self:
            if not wiz.data:
                raise exceptions.Warning(_("You need to select a file!"))
            date_hour = fields.datetime.now()
            actual_date = fields.date.today()
            filename = wiz.name
            if wiz.file_type == 'csv':
                counter = self._import_csv1d(template.id, wiz.data, wiz.delimeter)
            elif wiz.file_type == 'xls':
                counter = self._import_xls1d(template.id, wiz.data)
            elif wiz.file_type == 'xlsx':
                counter = self._import_xlsx1d(template.id, wiz.data)
            else:
                raise exceptions.Warning(_("Not a .csv/.xls/.xlsx file found"))
        return counter

